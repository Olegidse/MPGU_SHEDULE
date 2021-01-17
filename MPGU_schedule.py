import csv
import pandas as pd
import math as m
import re
from datetime import datetime, date, time
import xlrd
import openpyxl
from openpyxl.utils import range_boundaries
start_of_sch = 13
year = '2020'
noplace_str = 'Место занятия не указано'
schedule_file_name = 'Bakalavriat_3_k_5_s_20-21_Ochnoe (1).xlsx'
def isNaN(string):
    return string != string
def format_start_time(str):
    time = str
    start_time = time.split('-')[0]
    if len(start_time) < 5:
        start_time = '0' + start_time
    start_time = start_time[0:2]+ ':' + start_time[3:len(start_time)]
    start_time = datetime.strptime(start_time, "%H:%M")
    start_time = start_time.strftime("%I:%M %p")
    return start_time

def format_end_time(str):
    time = str
   
    end_time = time.split('-')[1]
    if len(end_time) < 5:
        end_time = '0' + end_time
    end_time = end_time[0:2]+ ':' + end_time[3:len(end_time)]
    end_time = datetime.strptime(end_time, "%H:%M")
    end_time = end_time.strftime("%I:%M %p")
    return end_time

def divide_subject(str):
    subj_list = []
    split_str = re.findall('\s{4,}',str)
    if len(split_str) == 0:
        new_list = [str]
    else:
        new_list = []
    i =0
    while i < len(split_str):
        lst = str.split(split_str[i])
        new_list.append(lst[0])
        str = lst[1]
        if i+1 == len(split_str):
            new_list.append(lst[1])  
        i+=1
        
    
    for item in new_list:
        places = find_place(item, True)
        new_list_2 = item.split('+')
        if len(new_list_2)>1:
            if len(places)==2:
                new_list_2[0] = new_list_2[0].replace(places[0],'')
                new_list_2[0] = new_list_2[0].replace(places[1],'')
                new_list_2[0] = new_list_2[0] + ' ' + places[0] + ' ' + places[1]
                subj_list.append(new_list_2[0])
                
                new_list_2[1] = new_list_2[1].replace(places[0],'')
                new_list_2[1] = new_list_2[1].replace(places[1],'')
                new_list_2[1] = new_list_2[1] + ' ' + places[0] + ' ' + places[1]
                subj_list.append(make_str_plus(new_list_2[0])+new_list_2[1])
            elif len(places) == 1:
                new_list_2[0] = new_list_2[0].replace(places[0],'')
                new_list_2[0] = new_list_2[0] + ' ' + places[0] 
                subj_list.append(new_list_2[0])

                new_list_2[1] = new_list_2[1].replace(places[0],'')
                new_list_2[1] = new_list_2[1] + ' ' + places[0] 
                subj_list.append(make_str_plus(new_list_2[0])+new_list_2[1])
            else:
                subj_list.append(new_list_2[0])
                subj_list.append(make_str_plus(new_list_2[0])+new_list_2[1])
           
        else:
            subj_list.append(new_list_2[0])
       
    return subj_list

def find_name(str):
    matches = re.findall('.{1,}\(\w{1,}\.\d{1,}\)', str)
    matches2 = re.findall('.{1,}\(\s{0,}\d', str)
    print(matches)
    print(matches2)
    if len(matches) > 0 and len(matches2) == 0:
        return ('"' + matches[0][0:len(matches[0])-9]+ '"')
    elif len(matches) == 0 and len(matches2) > 0:
            return ('"'+ matches2[0][0:len(matches2[0])-3]+'"')
    elif len(matches) > 0 and len(matches2) > 0:
        if len(matches[0]) < len(matches2[0]):
            return ('"' + matches[0][0:len(matches[0])-9]+ '"')
        else: 
            return ('"'+ matches2[0][0:len(matches2[0])-3]+'"')
    else:
        return False

def make_str_plus(str):
    matches = re.findall('.{1,}\(\w\w\)', str)
    part1 = matches[0][0:len(matches[0]) - 4]
    matches = re.findall('\(\w\w\).{1,}\(\d\d\.', str)
    part2 = matches[0][4:len(matches[0])-4]
    return part1+part2


def find_place(str, divide):
    matches = re.findall('\(ауд\.\d{1,}\)',str)
    matches2 = re.findall('https://\S+',str)
    if len(matches) > 0 and len(matches2) == 0: 
        return matches[0]
    elif len(matches) > 0 and len(matches2)> 0:
       if divide == False:    
            return matches[0] + " " + matches2[0]
       elif divide == True:
           list_of_places = []
           list_of_places.append(matches[0])
           list_of_places.append(matches2[0])
           return list_of_places
    elif len(matches) == 0 and len(matches2) > 0:
        return matches2[0]
    else:
        return noplace_str
        
       

def find_date(str):
    matches = re.findall('\d\d\.\d\d',str)   #\([\d\d\.\d\d, ]{1,}\)
    matches2 = re.findall('\D\d\.\d\d.',str)
    if len(matches2) > 0:
        for date in matches2:
            matches.append('0'+date[1:len(date)])
    if len(matches) > 0:
        return matches
    else:
        return False

def find_week_days(xlsx_name,col_index):
    df = pd.read_excel(xlsx_name)
    i = 0
    list_of_days = []
    flag = False
    while i < df.shape[0]:
       if  isNaN(df.loc[i][col_index]) == False:
           if len(re.findall('понедельник|вторник|среда|четверг|пятница|суббота',df.loc[i][col_index])) > 0:
               j = i+1
               flag2 = False
               while j < df.shape[0] and flag2 == False:
                   if j == df.shape[0] - 1:
                       list_of_days.append(range(i,j+1))
                       flag2 = True
                   if  isNaN(df.loc[j][col_index]) == False:
                        if len(re.findall('понедельник|вторник|среда|четверг|пятница|суббота',df.loc[j][col_index])) > 0:
                  
                            list_of_days.append(range(i,j))
                            i = j - 1 
                            flag2 = True
                   j+=1
       i+=1
                    
    return list_of_days


def read_xlsx(xlsx_name, col_index, time_index, week_days):
    df = pd.read_excel(xlsx_name)
    i = start_of_sch
    csv_list = [['Subject','Description', 'Start Date','End Date','Start Time', 'End Time', 'Location']]
    mistake_counter = 0  
    while i < df.shape[0]:
            
       if  isNaN(df.loc[i][col_index]) == False:
            #csv_list.append(df.loc[i][col_index])
            subject = df.loc[i][col_index]
            subject_list = divide_subject(subject)
            for item in subject_list:
              if not find_date(item) == False:
                dates = find_date(item)
                #dates = re.split(', +', dates)
                
                for j in range(len(dates)):
                    rows =[]
                    
                    rows.append(find_name(item))
                    
                    d = date(2020,int(dates[j][3:5]),int(dates[j][0:2]))
                    day = 0
                    error = 0
                    for k in range(len(week_days)):
                        if i in week_days[k]:
                            day = k
                            #print(week_days[day])
                            #print(i)
                    if not d.weekday() == day:
                        mistake_counter += 1
                        #print('Не совпадает:', dates[j][3:5] + '/'+dates[j][0:2]+ '/'+year, 'По факту: ',d.weekday(),"В расписании: ", day)
                        error = day - d.weekday()
                        if error == -5:
                            error = 2
                        if error == 5:
                            error = -2

                        rows.append('Дата в оригинальном расписании поставлена неверно, поэтому была изменена! Возможно эта херня вообще в другом месяце! Чтобы быть уверенными на 100%, обратитесь к знающим людям или к гадалке :)')
                    else:
                        rows.append('')
                    rows.append(dates[j][3:5] + '/'+str(int(dates[j][0:2])+error)+ '/'+year) #месяц+число+год
                    rows.append(dates[j][3:5] + '/'+str(int(dates[j][0:2])+error)+ '/'+year)
                    if  isNaN(df.loc[i][time_index]) == False:
                         rows.append(format_start_time (df.loc[i][time_index]))
                         rows.append(format_end_time (df.loc[i][time_index]))
                      
                    else:
                        rows.append(format_start_time (df.loc[i-1][time_index]))
                        rows.append(format_end_time (df.loc[i-1][time_index]))
                    rows.append(find_place(item, False))
                    print(rows)
                    csv_list.append(rows)
       i+=1
    
    print(mistake_counter)
    return csv_list

def find_merged(excel,letter):
    wb = openpyxl.load_workbook(excel)
    ws = wb[wb.get_sheet_names()[0]]
    merged = ws.merged_cell_ranges
    merged_cells = []
    for merge in merged:
        print(merge)
        range = str(merge)
        if range.split(':')[0]< letter < range.split(':')[1]:
            print(ws[range.split(':')[0]].value)

#find_merged('Bakalavriat_3_k_5_s_20-21_Ochnoe (1).xlsx', 'U')

def unmerge_all(file,new_name):
    counter = 0
    wbook=openpyxl.load_workbook(file)
    sheet=wbook[wbook.get_sheet_names()[0]]
    merged_ranges = [] 
    for group in sheet.merged_cells.ranges:
        merged_ranges.append(group)
    for cell_group in merged_ranges:
        min_col, min_row, max_col, max_row = cell_group.bounds
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value  
    wbook.save(new_name)



week_days = find_week_days(schedule_file_name, 1)   
unmerge_all(schedule_file_name,'unmerged.xlsx')  
csv_list = read_xlsx('unmerged.xlsx', 20, 3, week_days)
myFile = open('Schedule5.csv', 'w')
with myFile:
    writer = csv.writer(myFile, delimiter=',')
    writer.writerows(csv_list)
  

#unmerge_all("unmerged.xlsx")